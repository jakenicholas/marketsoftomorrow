#!/usr/bin/env python3
"""
One-shot Monday.com → Studio campaign migration.

Reads the 15 campaign rows from Campaigns_*.xlsx (Live + Completed sections),
creates one campaign per row, optionally parses "$X,XXX/mo" out of the Notes
column to populate monthly_income, then auto-links existing published posts
that mention the campaign name within its timeline window.

For each campaign:
  1. Parse monthly_income from notes (e.g. "$4,500/mo starting March '25").
  2. POST /campaigns to create.
  3. Resolve project_slug from projects-flat.json (campaign name → project).
     Stored on the campaign so the dashboard can group posts by project.
  4. Search /posts?q=<name>&status=published, filter to posts within
     ±LINK_WINDOW_DAYS of the campaign's start/end window, link each via
     POST /campaigns/:id/link-post (worker auto-fills post.income from
     campaign math).

Run:
    python3 scripts/migrate_campaigns.py --dry-run    # preview
    python3 scripts/migrate_campaigns.py              # live
"""
import argparse
import datetime
import json
import os
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request

import openpyxl

WORKER = 'https://tmw.jake-ab7.workers.dev'
XLSX   = os.environ.get('CAMPAIGNS_XLSX', '/Users/jakenicholas/Downloads/Campaigns_1781707726.xlsx')
PROJECTS_FLAT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'projects-flat.json')

# How far outside the campaign window a post can land and still be linked.
# Some posts publish a few days early ("preview" announcements) or late
# (recap pieces); 14 days catches both without dragging in unrelated content.
LINK_WINDOW_DAYS = 14

# Monday tier strings map straight through, but normalize casing.
TIER_MAP = {
    'gold': 'Gold', 'platinum': 'Platinum', 'custom': 'Custom',
}

STOP_WORDS = {
    'the','a','an','of','and','in','at','to','for','on','with','by','from',
    'is','it','its','as','that','this','their','our','your',
}

HTTP_PAUSE_S = 0.04


# ─── Helpers ───────────────────────────────────────────────────────────────
def tokens(s):
    s = re.sub(r"[^a-z0-9 ]", " ", (s or '').lower())
    return [t for t in s.split() if t and t not in STOP_WORDS and len(t) > 1]

def get_admin_token():
    r = subprocess.run(
        ['security', 'find-generic-password', '-s', 'tmw-admin-token', '-a', 'jakenicholas', '-w'],
        capture_output=True, text=True
    )
    tok = r.stdout.strip()
    if not tok:
        raise RuntimeError("No tmw-admin-token in keychain (run /studio/ login first)")
    return tok

TOKEN = None

def api(method, path, body=None, retries=2):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(
        WORKER + path,
        data=data, method=method,
        headers={
            'Authorization': f'Bearer {TOKEN}',
            'Content-Type': 'application/json',
            'User-Agent': 'tmw-campaigns-migration/1.0',
        },
    )
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError as e:
            txt = ''
            try: txt = e.read().decode('utf-8', errors='replace')
            except Exception: pass
            if e.code in (502, 503, 504) and attempt < retries:
                time.sleep(0.5 * (attempt + 1)); continue
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {txt[:300]}")
        except Exception:
            if attempt < retries:
                time.sleep(0.5 * (attempt + 1)); continue
            raise


# ─── Excel reading ─────────────────────────────────────────────────────────
def read_campaigns(path):
    """Read the Live + Completed campaign rows.

    Columns: A=Name, B=Start, C=End, D=Status (Live/Ended), E=Tier,
             F=Income (total), G=Notes, H=Item ID."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['campaigns']
    out = []
    for r in range(4, ws.max_row + 1):
        name = ws.cell(row=r, column=1).value
        if not name or name in ('Name', 'Subitems', 'Campaigns', 'Live', 'Completed'):
            continue
        status_raw = ws.cell(row=r, column=4).value
        if not status_raw or status_raw == 'Status':
            continue
        out.append({
            'row_n':       r,
            'name':        str(name).strip(),
            'start_date':  ws.cell(row=r, column=2).value,
            'end_date':    ws.cell(row=r, column=3).value,
            'status_raw':  str(status_raw),
            'tier':        ws.cell(row=r, column=5).value,
            'income':      ws.cell(row=r, column=6).value,
            'notes':       (ws.cell(row=r, column=7).value or '').strip() if ws.cell(row=r, column=7).value else '',
            'monday_id':   str(ws.cell(row=r, column=8).value or '').strip(),
        })
    return out


def parse_monthly_from_notes(notes):
    """'$4,500/mo starting March 25' → 4500.0
       Returns None if no match."""
    if not notes: return None
    m = re.search(r'\$\s*([\d,]+(?:\.\d+)?)\s*/\s*mo', notes, re.IGNORECASE)
    if not m: return None
    try:
        return float(m.group(1).replace(',', ''))
    except ValueError:
        return None


# ─── Project resolution ───────────────────────────────────────────────────
def build_project_index():
    with open(PROJECTS_FLAT, 'r', encoding='utf-8') as f:
        projects = json.load(f)
    by_title = {}
    by_tokens = []
    for p in projects:
        title = (p.get('Title') or '').strip()
        slug  = (p.get('Slug')  or '').strip()
        if not title or not slug: continue
        by_title[title.lower()] = slug
        by_tokens.append((set(tokens(title)), slug, title))
    return by_title, by_tokens


def resolve_project_slug(name, by_title, by_tokens):
    """Same containment-first matching as the monday-posts migration."""
    if not name: return None
    n_lc = name.lower().strip()
    if n_lc in by_title: return by_title[n_lc]
    nt = set(tokens(name))
    if not nt: return None
    best, best_score = None, 0.0
    for tt, slug, title in by_tokens:
        if not tt: continue
        if n_lc in title.lower(): score = 0.95
        elif all(tok in tt for tok in nt): score = 0.80
        else: score = len(nt & tt) / max(1, len(nt | tt))
        if score > best_score: best, best_score = slug, score
    return best if best_score >= 0.55 else None


# ─── Post-linking ──────────────────────────────────────────────────────────
def find_posts_for_campaign(name, start_ts, end_ts):
    """Worker scans title+excerpt+body for every brand token; filter to those
       within ±LINK_WINDOW_DAYS of the campaign window."""
    if not name: return []
    try:
        d = api('GET', '/posts?status=published&limit=100&q=' + urllib.parse.quote(name))
    except Exception:
        return []
    win_start = (start_ts or 0) - LINK_WINDOW_DAYS * 86400
    # Ongoing campaigns (no end_date): allow any post published from start
    # forward, capped at "today + window" so we don't grab future drafts that
    # somehow have a forward published_at.
    if end_ts:
        win_end = end_ts + LINK_WINDOW_DAYS * 86400
    else:
        win_end = int(time.time()) + LINK_WINDOW_DAYS * 86400
    hits = []
    for stub in (d.get('items') or []):
        pub = stub.get('published_at')
        if not pub: continue
        if pub < win_start or pub > win_end: continue
        hits.append(stub)
    return hits


# ─── Main ──────────────────────────────────────────────────────────────────
def main():
    global TOKEN
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help="preview only — no writes")
    args = parser.parse_args()
    TOKEN = get_admin_token()

    print(f"Reading campaigns from {XLSX}...")
    rows = read_campaigns(XLSX)
    print(f"  ✓ {len(rows)} campaign rows")

    print("Loading project index...")
    by_title, by_tokens_p = build_project_index()
    print(f"  ✓ {len(by_title)} projects indexed")

    stats = {'created': 0, 'linked': 0, 'skipped': 0}
    summary = []

    for row in rows:
        name = row['name']
        status = 'ended' if row['status_raw'].lower() == 'ended' else 'live'
        tier = TIER_MAP.get((row['tier'] or '').lower(), row['tier'])
        # Drop tier values that aren't real ("Campaign" was the heading row's
        # value that snuck through).
        if tier in ('Campaign', None, ''): tier = None
        start_ts = int(row['start_date'].timestamp()) if isinstance(row['start_date'], datetime.datetime) else None
        end_ts   = int(row['end_date']  .timestamp()) if isinstance(row['end_date']  , datetime.datetime) else None
        total_income   = None
        if row['income'] not in (None, ''):
            try: total_income = float(row['income'])
            except (TypeError, ValueError): pass
        monthly = parse_monthly_from_notes(row['notes'])
        project_slug = resolve_project_slug(name, by_title, by_tokens_p)

        payload = {
            'name':           name,
            'status':         status,
            'tier':           tier,
            'start_date':     start_ts,
            'end_date':       end_ts,
            'total_income':   total_income,
            'monthly_income': monthly,
            'notes':          row['notes'] or None,
            'project_slug':   project_slug,
        }

        # ── Create
        if args.dry_run:
            cid = f'DRY-{stats["created"]+1}'
        else:
            try:
                d = api('POST', '/campaigns', payload)
                cid = d['campaign']['id']
                time.sleep(HTTP_PAUSE_S)
            except RuntimeError as e:
                print(f"  ✗ create failed for {name!r}: {e}")
                stats['skipped'] += 1
                continue
        stats['created'] += 1

        # ── Link posts
        post_hits = find_posts_for_campaign(name, start_ts, end_ts)
        linked = 0
        for stub in post_hits:
            if args.dry_run:
                linked += 1; continue
            try:
                api('POST', f'/campaigns/{urllib.parse.quote(cid)}/link-post', {'post_id': stub['id']})
                linked += 1
                time.sleep(HTTP_PAUSE_S)
            except RuntimeError as e:
                print(f"    ✗ link failed for post {stub.get('slug','?')}: {e}")
        stats['linked'] += linked

        summary.append({
            'name': name,
            'cid': cid,
            'status': status,
            'tier': tier,
            'income': total_income,
            'monthly': monthly,
            'project_slug': project_slug,
            'linked_posts': linked,
            'total_post_hits': len(post_hits),
        })
        marker = '🟢' if status == 'live' else '⚫'
        print(f"  {marker} {name:40s} → {linked}/{len(post_hits)} posts linked"
              + (f"  [{tier}]" if tier else '')
              + (f"  ${total_income:,.0f}" if total_income else '')
              + (f"  +${monthly:,.0f}/mo" if monthly else '')
              + (f"  proj={project_slug}" if project_slug else ''))

    print()
    print("── Summary " + "─" * 50)
    print(f"  Campaigns created:  {stats['created']}")
    print(f"  Posts linked:       {stats['linked']}")
    print(f"  Skipped:            {stats['skipped']}")
    if args.dry_run:
        print("  (dry-run — no writes)")


if __name__ == '__main__':
    main()
