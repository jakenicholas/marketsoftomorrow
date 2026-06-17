#!/usr/bin/env python3
"""
One-shot Monday.com → Studio migration.

Reads every 'Posted' row from the Monday.com export (Paid_Editorial_*.xlsx)
and backfills post_type / income / contact_id / project_slug on the matching
live posts in D1.

For each row:
  1. Resolve / create the contact (name + email + company from cols K/L/M).
     Re-uses an existing contact when the name OR email already matches one;
     never duplicates.
  2. Find the live post by  (a) published_at within ±DATE_WINDOW_DAYS of the
     row's Launch Date  AND  (b) the highest title-Jaccard score against
     (brand + row-name) tokens. Threshold MATCH_MIN.
  3. Resolve the project_slug from the Brand column against
     projects-flat.json (canonical title match, then fuzzy).
  4. PATCH the post with post_type, income, contact_id, project_slug.

Run:
    python3 scripts/migrate_monday.py --dry-run     # preview, no writes
    python3 scripts/migrate_monday.py               # live writes
    python3 scripts/migrate_monday.py --limit 20    # do the first 20 rows
"""
import argparse
import datetime
import json
import os
import re
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request

import openpyxl

WORKER = 'https://tmw.jake-ab7.workers.dev'
XLSX   = os.environ.get('MIGRATE_XLSX', '/Users/jakenicholas/Downloads/Paid_Editorial_1781701593.xlsx')
PROJECTS_FLAT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'projects-flat.json')

DATE_WINDOW_DAYS = 10
MATCH_MIN        = 0.18   # min Jaccard-weighted score to accept a post match
HTTP_PAUSE_S     = 0.04   # gentle pacing on the worker (~25 req/s)

# Accept the five canonical post types; map a handful of stray Monday values
# to sane defaults. Anything else falls through to 'Editorial'.
POST_TYPE_MAP = {
    'editorial': 'Editorial',
    'paid': 'Paid',
    'barter': 'Barter',
    'potential barter': 'Potential Barter',
    'partner': 'Partner',
    # Noise rows in the sheet — preserve as Editorial unless we know better.
    'video reel': 'Editorial',
    'sculpture': 'Editorial',
    'sept update': 'Editorial',
    'awaiting payment': 'Paid',   # one row had income captured — treat as paid
}

STOP_WORDS = {
    'the','a','an','of','and','in','at','to','for','on','with','by','from',
    'is','it','its','as','that','this','their','our','your','de','la','le','el',
}


# ─── Token helpers ─────────────────────────────────────────────────────────
def tokens(s):
    s = re.sub(r"[^a-z0-9 ]", " ", (s or '').lower())
    return [t for t in s.split() if t and t not in STOP_WORDS and len(t) > 1]

def jaccard(a, b):
    sa, sb = set(a), set(b)
    if not sa or not sb: return 0.0
    return len(sa & sb) / len(sa | sb)


# ─── Auth ──────────────────────────────────────────────────────────────────
def get_admin_token():
    """Pull the admin token from macOS keychain (set by /studio/ login)."""
    r = subprocess.run(
        ['security', 'find-generic-password', '-s', 'tmw-admin-token', '-a', 'jakenicholas', '-w'],
        capture_output=True, text=True
    )
    tok = r.stdout.strip()
    if not tok:
        raise RuntimeError("No tmw-admin-token in keychain (run /studio/ login first)")
    return tok

TOKEN = None  # populated in main()


def api(method, path, body=None, retries=2):
    data = json.dumps(body).encode() if body is not None else None
    req = urllib.request.Request(
        WORKER + path,
        data=data, method=method,
        headers={
            'Authorization': f'Bearer {TOKEN}',
            'Content-Type': 'application/json',
            'User-Agent': 'tmw-monday-migration/1.0',
        },
    )
    for attempt in range(retries + 1):
        try:
            with urllib.request.urlopen(req, timeout=30) as r:
                return json.loads(r.read())
        except urllib.error.HTTPError as e:
            txt = ''
            try: txt = e.read().decode('utf-8', errors='replace')
            except Exception: pass
            if e.code in (502, 503, 504) and attempt < retries:
                time.sleep(0.5 * (attempt + 1))
                continue
            raise RuntimeError(f"HTTP {e.code} {method} {path}: {txt[:300]}")
        except Exception as e:
            if attempt < retries:
                time.sleep(0.5 * (attempt + 1))
                continue
            raise


# ─── Excel reading ─────────────────────────────────────────────────────────
def read_posted_rows(path):
    """Return [{row_n, name, launch_date, post_type, brand, contact_name,
              email, company, income, monday_id}, ...] for Status='Posted'."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['paid & editorial']
    rows = []
    for r in range(6, ws.max_row + 1):
        if ws.cell(row=r, column=5).value != 'Posted':
            continue
        rows.append({
            'row_n':        r,
            'name':         str(ws.cell(row=r, column=1).value or '').strip(),
            'launch_date':  ws.cell(row=r, column=4).value,
            'post_type':    str(ws.cell(row=r, column=7).value or '').strip(),
            'brand':        str(ws.cell(row=r, column=10).value or '').strip(),
            'contact_name': str(ws.cell(row=r, column=11).value or '').strip(),
            'email':        str(ws.cell(row=r, column=12).value or '').strip(),
            'company':      str(ws.cell(row=r, column=13).value or '').strip(),
            'income':       ws.cell(row=r, column=14).value,
            'monday_id':    str(ws.cell(row=r, column=17).value or '').strip(),
        })
    return rows


# ─── Project index (brand → slug) ──────────────────────────────────────────
def build_project_index():
    """{normalized_title: slug} from projects-flat.json, plus a token list
       per project for the fuzzy fallback."""
    with open(PROJECTS_FLAT, 'r', encoding='utf-8') as f:
        projects = json.load(f)
    by_title = {}
    by_tokens = []   # [(token_set, slug, title)]
    for p in projects:
        title = (p.get('Title') or '').strip()
        slug  = (p.get('Slug') or '').strip()
        if not title or not slug: continue
        by_title[title.lower()] = slug
        by_tokens.append((set(tokens(title)), slug, title))
    return by_title, by_tokens


def resolve_project_slug(brand, by_title, by_tokens):
    if not brand: return None, None
    b = brand.lower().strip()
    if b in by_title: return by_title[b], 'exact'
    bt = set(tokens(brand))
    if not bt: return None, None
    best, best_score = None, 0.0
    for tt, slug, title in by_tokens:
        if not tt: continue
        # Substring containment: the brand string appears verbatim in the
        # project title (handles "Pendry Tampa" → "The Pendry Tampa and
        # Residences") — strongest signal, jump straight to a high score.
        if b in title.lower():
            score = 0.95
        # Every brand token is in the project's tokens (any order). Catches
        # "Marina Pointe" → "Marina Pointe: Luna" where the brand is a strict
        # subset of the longer canonical name.
        elif all(tok in tt for tok in bt):
            score = 0.80
        else:
            score = len(bt & tt) / max(1, len(bt | tt))
        if score > best_score:
            best, best_score = (slug, title), score
    if best and best_score >= 0.55:
        return best[0], f'match:{best_score:.2f}'
    return None, None


# ─── Post matching ─────────────────────────────────────────────────────────
def load_all_posts():
    """Pull every post (published + draft) into one list with the fields we
       need for matching: id, slug, title, published_at."""
    all_posts = []
    for status in ('published', 'draft'):
        offset, total = 0, 1
        while offset < total:
            data = api('GET', f'/posts?limit=1500&offset={offset}&status={status}')
            items = data.get('items', [])
            total = data.get('total', 0)
            for p in items:
                all_posts.append({
                    'id': p['id'], 'slug': p['slug'], 'title': p['title'],
                    'published_at': p.get('published_at'),
                    'post_type': p.get('post_type'),
                    'contact_id': p.get('contact_id'),
                })
            if not items: break
            offset += len(items)
    return all_posts


def index_posts_by_date(posts):
    """Bucket posts by their published-day so date-windowed lookup is O(window
       * avg-bucket) instead of O(n)."""
    by_day = {}
    for p in posts:
        if not p['published_at']: continue
        d = datetime.date.fromtimestamp(p['published_at'])
        by_day.setdefault(d, []).append(p)
    return by_day


def find_post(row, by_day):
    if not isinstance(row['launch_date'], datetime.datetime):
        return None, 0.0, 'no-launch-date'
    target = row['launch_date'].date()
    candidates = []
    for delta in range(-DATE_WINDOW_DAYS, DATE_WINDOW_DAYS + 1):
        d = target + datetime.timedelta(days=delta)
        candidates.extend(by_day.get(d, []))
    if not candidates:
        return None, 0.0, 'no-posts-in-window'
    brand    = (row['brand'] or '').strip()
    brand_lc = brand.lower()
    brand_t  = tokens(brand)
    # Strip the date suffix off the row name (e.g. "South Flagler House: March"
    # → "South Flagler House") so the campaign month doesn't pollute matching.
    name_no_suffix = row['name'].split(':')[0]
    name_t = tokens(name_no_suffix)
    best, best_score = None, 0.0
    for p in candidates:
        title    = (p['title'] or '').strip()
        title_lc = title.lower()
        tt       = tokens(title)
        # ── Containment bonuses: the strongest possible signal.
        contain = 0.0
        if brand_lc and brand_lc in title_lc:
            contain += 0.6                                   # full brand string appears verbatim
        elif brand_t and all(tok in tt for tok in brand_t):
            contain += 0.45                                  # every brand token is in the title (any order)
        # ── Jaccard backbone
        jb = jaccard(brand_t, tt)
        jn = jaccard(name_t,  tt)
        s = contain + 2 * jb + jn
        # Date-distance tiebreak (closer wins ties).
        dd = abs((datetime.date.fromtimestamp(p['published_at']) - target).days)
        s -= 0.005 * dd
        if s > best_score:
            best, best_score = p, s
    if best and best_score >= MATCH_MIN:
        return best, best_score, ''
    return None, best_score, 'below-threshold'


# ─── Contact resolution ────────────────────────────────────────────────────
def normalize_email(s): return (s or '').strip().lower()
def normalize_name(s):  return ' '.join((s or '').strip().lower().split())


def load_existing_contacts():
    """Pull all existing contacts so we can match without re-creating dups."""
    contacts = []
    offset, total = 0, 1
    while offset < total:
        d = api('GET', f'/contacts?limit=1000&offset={offset}')
        items = d.get('items', [])
        total = d.get('total', 0)
        contacts.extend(items)
        if not items: break
        offset += len(items)
    by_email = {}
    by_name  = {}
    for c in contacts:
        if c.get('email'): by_email[normalize_email(c['email'])] = c
        if c.get('name'):  by_name.setdefault(normalize_name(c['name']), c)
    return contacts, by_email, by_name


# ─── Main ──────────────────────────────────────────────────────────────────
def main():
    global TOKEN
    parser = argparse.ArgumentParser()
    parser.add_argument('--dry-run', action='store_true', help="don't write — preview only")
    parser.add_argument('--limit',   type=int, default=0,  help="cap rows processed")
    parser.add_argument('--start',   type=int, default=0,  help="skip the first N rows")
    parser.add_argument('--unmatched-csv', default='monday_unmatched.csv', help='CSV of unmatched rows')
    args = parser.parse_args()

    TOKEN = get_admin_token()

    print(f"Reading rows from {XLSX}...")
    rows = read_posted_rows(XLSX)
    print(f"  ✓ {len(rows)} 'Posted' rows")

    print("Loading project index from projects-flat.json...")
    by_title, by_tokens_proj = build_project_index()
    print(f"  ✓ {len(by_title)} projects indexed")

    print("Loading all live posts from D1...")
    all_posts = load_all_posts()
    by_day = index_posts_by_date(all_posts)
    print(f"  ✓ {len(all_posts)} posts ({sum(len(v) for v in by_day.values())} dated)")

    print("Loading existing contacts...")
    _, by_email, by_name = load_existing_contacts()
    print(f"  ✓ {len(by_email)} contacts indexed by email, {len(by_name)} by name")

    # Pre-pass: resolve every (contact_name, email, company) into a contact id,
    # creating once per row group so repeats reuse the same id.
    contact_cache = {}  # key=(name_lower, email_lower) → contact dict
    project_cache = {}  # brand_lower → (slug, reason)

    matched = []
    unmatched = []
    stats = {'created_contacts': 0, 'reused_contacts': 0, 'updates': 0, 'project_links': 0}

    end = len(rows) if args.limit == 0 else min(len(rows), args.start + args.limit)
    for i, row in enumerate(rows[args.start:end], start=args.start):
        # ── post-type normalization
        pt_lc = (row['post_type'] or '').strip().lower()
        post_type = POST_TYPE_MAP.get(pt_lc, 'Editorial')

        # ── contact resolution
        contact = None
        if row['contact_name'] or row['email']:
            key = (normalize_name(row['contact_name']), normalize_email(row['email']))
            if key in contact_cache:
                contact = contact_cache[key]
            else:
                # Match strategy: prefer email exact match; else name exact match.
                email_lc = normalize_email(row['email'])
                if email_lc and email_lc in by_email:
                    contact = by_email[email_lc]
                    stats['reused_contacts'] += 1
                else:
                    name_lc = normalize_name(row['contact_name'])
                    if name_lc and name_lc in by_name:
                        contact = by_name[name_lc]
                        stats['reused_contacts'] += 1
                if contact is None and (row['contact_name'] or row['email']):
                    # Create a new one.
                    payload = {
                        'name':    row['contact_name'] or row['email'].split('@')[0],
                        'email':   row['email'] or None,
                        'company': row['company'] or None,
                    }
                    if args.dry_run:
                        contact = { 'id': f'DRY-{len(contact_cache)+1}', **payload, 'tags': [] }
                    else:
                        d = api('POST', '/contacts', payload)
                        contact = d['contact']
                        if contact.get('email'): by_email[normalize_email(contact['email'])] = contact
                        if contact.get('name'):  by_name.setdefault(normalize_name(contact['name']), contact)
                        time.sleep(HTTP_PAUSE_S)
                    stats['created_contacts'] += 1
                contact_cache[key] = contact

        # ── project_slug resolution from Brand
        slug = None
        if row['brand']:
            b_lc = row['brand'].lower().strip()
            if b_lc in project_cache:
                slug, _ = project_cache[b_lc]
            else:
                slug, reason = resolve_project_slug(row['brand'], by_title, by_tokens_proj)
                project_cache[b_lc] = (slug, reason)
            if slug:
                stats['project_links'] += 1

        # ── find matching post
        post, score, reason = find_post(row, by_day)
        if not post:
            unmatched.append({**row, 'reason': reason, 'best_score': round(score, 3)})
            continue

        # ── build PATCH payload (only fields we actually want to set)
        patch = {'post_type': post_type}
        if row['income'] not in (None, ''):
            try: patch['income'] = float(row['income'])
            except (TypeError, ValueError): pass
        if contact: patch['contact_id'] = contact['id']
        if slug:    patch['project_slug'] = slug

        matched.append({
            'monday_id': row['monday_id'],
            'monday_name': row['name'],
            'post_slug':  post['slug'],
            'post_title': post['title'],
            'score':      round(score, 3),
            'patch':      patch,
        })

        if not args.dry_run:
            api('PATCH', f"/posts/{urllib.parse.quote(post['id'])}", patch)
            time.sleep(HTTP_PAUSE_S)
        stats['updates'] += 1

        # Progress every 50
        if (i + 1) % 50 == 0:
            print(f"  [{i+1}/{end}] matched {len(matched)}, unmatched {len(unmatched)}")

    print()
    print("── Summary " + "─" * 50)
    print(f"  Posted rows considered: {end - args.start}")
    print(f"  Matched & {'(would) ' if args.dry_run else ''}updated posts: {len(matched)}")
    print(f"  Unmatched rows:        {len(unmatched)}")
    print(f"  Contacts created:      {stats['created_contacts']}")
    print(f"  Contacts reused:       {stats['reused_contacts']}")
    print(f"  Project links resolved: {stats['project_links']}")
    if args.dry_run:
        print("  (dry-run — no writes)")

    # Write unmatched CSV
    if unmatched:
        with open(args.unmatched_csv, 'w', encoding='utf-8') as f:
            f.write('row,monday_id,launch_date,post_type,brand,contact_name,email,best_score,reason\n')
            for u in unmatched:
                f.write(','.join([
                    str(u['row_n']),
                    f'"{u["monday_id"]}"',
                    str(u['launch_date'] or ''),
                    f'"{u["post_type"]}"',
                    f'"{u["brand"]}"',
                    f'"{u["contact_name"]}"',
                    f'"{u["email"]}"',
                    str(u.get('best_score', 0)),
                    f'"{u.get("reason","")}"',
                ]) + '\n')
        print(f"  Unmatched CSV: {args.unmatched_csv}")

    # Show first 8 matches as a sanity sample
    if matched:
        print("\nSample matches:")
        for m in matched[:8]:
            print(f"  {m['monday_id']}  '{m['monday_name'][:30]:30s}' → {m['post_slug'][:40]:40s}  score={m['score']}  {json.dumps(m['patch'])[:100]}")


if __name__ == '__main__':
    main()
